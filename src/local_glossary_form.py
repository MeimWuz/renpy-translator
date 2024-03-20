import os.path

import openpyxl
from PySide6.QtCore import QCoreApplication, Qt
from PySide6.QtGui import QStandardItemModel, QStandardItem
from PySide6.QtWidgets import QDialog, QTableWidget, QTableView, QHeaderView, QTableWidgetItem, QFileDialog, \
    QInputDialog, QMessageBox, QStyledItemDelegate, QPushButton
from openpyxl.workbook import Workbook

from local_glossary import Ui_LocalGlossaryDialog


class MyTableView(QTableView):
    def __init__(self, parent=None):
        super(MyTableView, self).__init__(parent)
        self.model = QStandardItemModel()
        self.setModel(self.model)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.model.setHorizontalHeaderLabels([QCoreApplication.translate('LocalGlossaryDialog', 'Original', None),
                                              QCoreApplication.translate('LocalGlossaryDialog', 'Replace', None)])
        self.model.setColumnCount(2)
        self.model.dataChanged.connect(self.handle_data_changed)
        self.file = None
        self.row = 0

    def handle_data_changed(self, top_left, bottom_right):
        self.file = self.selectFileText.toPlainText().replace('file:///', '')
        if len(self.file) == 0:
            return
        if not os.path.isfile(self.file):
            return
        model = self.model
        rows = model.rowCount()
        cols = model.columnCount()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=self.model.horizontalHeaderItem(0).text())
        ws.cell(row=1, column=2, value=self.model.horizontalHeaderItem(1).text())
        for r in range(rows):
            for c in range(cols):
                index = model.index(r, c)
                value = model.data(index, Qt.DisplayRole)
                if value is None:
                    continue
                ws.cell(row=r + 2, column=c + 1, value=value)
        wb.save(self.file)


class MyLocalGlossaryForm(QDialog, Ui_LocalGlossaryDialog):
    def __init__(self, parent=None):
        super(MyLocalGlossaryForm, self).__init__(parent)
        self.setupUi(self)
        self.tableView = MyTableView()
        self.tableView.selectFileText = self.selectFileText
        self.verticalLayout.addWidget(self.tableView)
        self.selectFileBtn.clicked.connect(self.select_file)
        self.selectFileText.textChanged.connect(self.on_text_changed)
        self.appendCheckBox.stateChanged.connect(self.on_append_checkbox_changed)
        self.data = None
        self.confirmButton.clicked.connect(self.on_confirm_clicked)
        self.create_file_param = None

    def on_append_checkbox_changed(self):
        self.load_from_xlsx(self.tableView.file)
        if self.appendCheckBox.isChecked():
            self.tableView.model.setRowCount((int(self.tableView.row / 100) + 1) * 100)
        else:
            self.tableView.model.setRowCount(self.tableView.row)

    def on_text_changed(self):
        file = self.selectFileText.toPlainText().replace('file:///', '')
        if len(file) == 0:
            self.tableView.model.clear()
            return
        if os.path.isfile(file) and file.endswith('.xlsx'):
            self.load_from_xlsx(file)
            self.tableView.file = file
            self.on_append_checkbox_changed()
        else:
            self.tableView.file = None
            self.tableView.model.clear()
            self.tableView.model.setHorizontalHeaderLabels(
                [QCoreApplication.translate('LocalGlossaryDialog', 'Original', None),
                 QCoreApplication.translate('LocalGlossaryDialog', 'Replace', None)])
            self.tableView.model.setColumnCount(1)
            self.tableView.model.setRowCount(1)
            self.tableView.row = 0
            if file.endswith('.xlsx'):
                text = file + ' : ' + QCoreApplication.translate(
                    'LocalGlossaryDialog', 'The file does not exist.Click to create it',
                    None)
                self.create_file_param = file
            else:
                if os.path.isfile(file + '.xlsx'):
                    text = file + ' : ' + QCoreApplication.translate(
                        'LocalGlossaryDialog', 'The file is not a xlsx file.Click to open ',
                        None) + file + '.xlsx'
                else:
                    text = file + ' : ' + QCoreApplication.translate(
                        'LocalGlossaryDialog', 'The file is not a xlsx file.Click to create ',
                        None) + file + '.xlsx'
                self.create_file_param = file + '.xlsx'
            m_button = QPushButton()
            m_button.setText(text)

            m_button.clicked.connect(self.create_file)
            self.tableView.setIndexWidget(self.tableView.model.index(0, 0), m_button)

    def create_file(self):
        if os.path.isfile(self.create_file_param):
            self.tableView.file = self.create_file_param
            self.selectFileText.setText(self.create_file_param)
            return
        model = self.tableView.model
        rows = model.rowCount()
        cols = model.columnCount()
        wb = Workbook()
        ws = wb.active
        self.tableView.model.clear()
        self.tableView.model.setHorizontalHeaderLabels(
            [QCoreApplication.translate('LocalGlossaryDialog', 'Original', None),
             QCoreApplication.translate('LocalGlossaryDialog', 'Replace', None)])
        ws.cell(row=1, column=1, value=self.tableView.model.horizontalHeaderItem(0).text())
        ws.cell(row=1, column=2, value=self.tableView.model.horizontalHeaderItem(1).text())
        self.tableView.file = self.create_file_param
        wb.save(self.tableView.file)
        # self.load_from_xlsx(self.tableView.file)
        self.selectFileText.setText(self.tableView.file)
        self.appendCheckBox.setChecked(True)

    def select_file(self):
        file, filetype = QFileDialog.getOpenFileName(self,
                                                     QCoreApplication.translate('LocalGlossaryDialog',
                                                                                'select the file you want to import',
                                                                                None),
                                                     '',
                                                     "Xlsx Files (*.xlsx)")
        if os.path.isfile(file):
            self.tableView.file = file
            self.selectFileText.setText(file)

    def load_from_xlsx(self, file):
        wb = openpyxl.load_workbook(file,data_only = True)
        ws = wb.active
        row = ws.max_row
        column = ws.max_column
        self.tableView.row = row
        self.tableView.column = column
        self.tableView.model.dataChanged.disconnect()
        if row > 0:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            first_row_list = list(first_row)
            self.tableView.model.clear()
            self.tableView.model.setHorizontalHeaderLabels(first_row_list)
            self.tableView.model.setRowCount(row)
            self.tableView.model.setColumnCount(column)
            for i in range(row):
                for j in range(column):
                    if i == 0:
                        continue
                    cell_value = ws.cell(row=i + 1, column=j + 1).value
                    if cell_value is None:
                        cell_value = ''
                    data = str(cell_value)
                    item = QStandardItem(data)
                    self.tableView.model.setItem(i-1, j, item)
        self.tableView.model.dataChanged.connect(self.tableView.handle_data_changed)
        wb.close()

    def get_data(self):
        dic = dict()
        model = self.tableView.model
        rows = model.rowCount()
        cols = model.columnCount()
        if cols < 2:
            return None
        for r in range(rows):
            index1 = model.index(r, 0)
            index2 = model.index(r, 1)
            value1 = model.data(index1, Qt.DisplayRole)
            value2 = model.data(index2, Qt.DisplayRole)
            if value1 is None or value2 is None:
                continue
            dic[value1] = value2
        return dic

    def on_confirm_clicked(self):
        self.data = self.get_data()
        self.close()
